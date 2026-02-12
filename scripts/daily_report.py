#!/usr/bin/env python3
"""
Daily submission report (IST) – with Day-by-Day Excel Matrix.

Outputs per run (for the IST date that just ended):
- reports/daily_report_YYYYMMDD.md
- reports/submitted_YYYYMMDD.csv
- reports/not_submitted_YYYYMMDD.csv
- reports/late_YYYYMMDD.csv
- reports/naming_issues_YYYYMMDD.csv
- NEW: reports/submissions_daily_matrix.xlsx
    Sheet "daily_counts":
      roll_number | full_name | github_username | 2026-02-01 | 2026-02-02 | ... | total_files | unique_days

Key behaviors:
- Scans student branches (local or 'origin/<username>') after 'git fetch --all --prune --tags'.
- Accepts multiple files per day:
    firstname_lastname_YYYYMMDD.py
    firstname_lastname_YYYYMMDD_1.py
    firstname_lastname_YYYYMMDD_2.py
- For 'late' marking uses the file’s last commit time vs 23:59:59 IST of that day.
- Rebuilds the Excel matrix from ALL reports/submitted_*.csv files each run (stateless & robust).

Environment variables (optional):
- TARGET_DATE_IST: YYYYMMDD or YYYY-MM-DD  (default: yesterday in IST)
- INCLUDE_MAIN_FALLBACK: "1" to also search 'main' if student branch is missing (default: off)
- MAIN_BRANCH_NAME: default "main"
- VERBOSE: "1" for extra diagnostic output
- ENFORCE_SINGLE_FILE: "1" to flag MULTIPLE_FILES as a naming issue (default: off)
"""

import os
import re
import csv
import sys
import glob
import subprocess
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

# Excel deps
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

IST = ZoneInfo("Asia/Kolkata")

SUBMISSIONS_DIR = "submissions"
REPORTS_DIR = "reports"
STUDENTS_CSV = "students.csv"

# Accept optional numeric index suffix after the date:  *_YYYYMMDD(_<n>)?.py
FILENAME_DATE_RE = re.compile(
    r"^(?P<prefix>.+)_(?P<yyyymmdd>\d{8})(?:_(?P<idx>\d+))?\.py$",
    re.IGNORECASE,
)

def env_flag(name: str, default: bool = False) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    return str(v).strip() in ("1", "true", "TRUE", "yes", "YES", "on", "On")

def env_str(name: str, default: str) -> str:
    v = os.getenv(name)
    return v if v is not None else default

def verbose(msg: str):
    if env_flag("VERBOSE", False):
        print(f"[VERBOSE] {msg}")

def run(cmd: list[str]) -> str:
    verbose(f"RUN: {' '.join(cmd)}")
    out = subprocess.check_output(cmd, stderr=subprocess.STDOUT)
    return out.decode("utf-8", errors="replace").strip()

def try_run(cmd: list[str]) -> tuple[bool, str]:
    try:
        return True, run(cmd)
    except subprocess.CalledProcessError as e:
        return False, e.output.decode("utf-8", errors="replace")

def ensure_dirs():
    os.makedirs(REPORTS_DIR, exist_ok=True)

def git_fetch_all():
    ok, out = try_run(["git", "fetch", "--all", "--prune", "--tags"])
    if not ok:
        print("WARNING: 'git fetch --all --prune --tags' failed. The script may miss remote branches.", file=sys.stderr)
        verbose(out)

def ref_exists(ref: str) -> bool:
    ok, _ = try_run(["git", "rev-parse", "--verify", "--quiet", f"{ref}^{{commit}}"])
    return ok

def preferred_ref_for_student(username: str) -> tuple[str | None, str]:
    local_ref = username
    remote_ref = f"origin/{username}"
    if ref_exists(local_ref):
        return local_ref, "origin"
    if ref_exists(remote_ref):
        return remote_ref, "origin"
    return None, "origin"

def list_branch_files(ref: str, path: str) -> list[str]:
    ok, out = try_run(["git", "ls-tree", "-r", "--name-only", ref, path])
    if not ok:
        return []
    files = [
        line.strip().split("/")[-1]
        for line in out.splitlines()
        if line.strip().lower().endswith(".py")
    ]
    return files

def get_last_commit_iso_for_path(ref: str, file_path: str) -> str | None:
    ok, out = try_run(["git", "log", "-1", "--format=%cI", ref, "--", file_path])
    if not ok:
        return None
    ts = out.strip()
    return ts or None

def parse_target_date_ist() -> datetime.date:
    override = os.getenv("TARGET_DATE_IST")
    if override:
        s = override.strip()
        try:
            if re.fullmatch(r"\d{8}", s):
                return datetime.strptime(s, "%Y%m%d").date()
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            print(f"WARNING: Invalid TARGET_DATE_IST={s!r}. Falling back to yesterday IST.", file=sys.stderr)
    now_ist = datetime.now(tz=IST)
    return (now_ist - timedelta(days=1)).date()

def read_students() -> list[dict]:
    students: list[dict] = []
    with open(STUDENTS_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if "github_username" not in reader.fieldnames:
            raise ValueError("students.csv must contain a 'github_username' column.")
        has_full = "full_name" in reader.fieldnames
        has_roll = "roll_number" in reader.fieldnames
        for row in reader:
            username = (row.get("github_username") or "").strip()
            if not username:
                continue
            students.append({
                "github_username": username,
                "full_name": (row.get("full_name") or "").strip() if has_full else "",
                "roll_number": (row.get("roll_number") or "").strip() if has_roll else "",
            })
    return students

def ist_deadline_for_date(d: datetime.date) -> datetime:
    return datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=IST)

def build_daily_counts_from_reports(reports_dir: str) -> tuple[list[str], dict[str, dict[str, int]]]:
    """
    Parse all 'submitted_YYYYMMDD.csv' under reports/ and build:
    - dates: sorted list of 'YYYY-MM-DD' strings
    - counts: dict[date_str][github_username] = files_count_on_that_date
    """
    pattern = os.path.join(reports_dir, "submitted_*.csv")
    date_to_counts: dict[str, dict[str, int]] = {}
    for path in sorted(glob.glob(pattern)):
        base = os.path.basename(path)
        # Expect: submitted_YYYYMMDD.csv
        m = re.match(r"submitted_(\d{8})\.csv$", base)
        if not m:
            continue
        yyyymmdd = m.group(1)
        date_str = datetime.strptime(yyyymmdd, "%Y%m%d").date().isoformat()  # YYYY-MM-DD
        per_user = date_to_counts.setdefault(date_str, {})
        try:
            with open(path, newline="", encoding="utf-8") as f:
                r = csv.DictReader(f)
                for row in r:
                    u = (row.get("github_username") or "").strip()
                    if not u:
                        continue
                    per_user[u] = per_user.get(u, 0) + 1  # count files
        except Exception as e:
            print(f"WARNING: Could not read {path}: {e}", file=sys.stderr)

    dates = sorted(date_to_counts.keys())
    return dates, date_to_counts

def write_excel_daily_matrix(path: str, students: list[dict], dates: list[str], date_to_counts: dict[str, dict[str, int]]):
    if not OPENPYXL_AVAILABLE:
        print("WARNING: openpyxl not available; skipping Excel matrix.", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "daily_counts"

    # Header row
    fixed_headers = ["roll_number", "full_name", "github_username"]
    headers = fixed_headers + dates + ["total_files", "unique_days"]
    ws.append(headers)

    # Sort student rows by roll_number if present, else by username
    def sort_key(s):
        rn = s.get("roll_number", "")
        return (rn if rn else "ZZZ", s["github_username"])

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    gray = PatternFill("solid", fgColor="EEEEEE")

    # Apply header styles
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        if header in ("total_files", "unique_days"):
            cell.fill = gray

    # Data rows
    for s in sorted(students, key=sort_key):
        u = s["github_username"]
        row_values = [
            s.get("roll_number", "") or "",
            s.get("full_name", "") or "",
            u,
        ]
        total = 0
        unique_days = 0
        for d in dates:
            c = date_to_counts.get(d, {}).get(u, 0)
            row_values.append(c)
            total += c
            if c > 0:
                unique_days += 1
        row_values += [total, unique_days]
        ws.append(row_values)

    # Simple autosize
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(path)
    print(f"Wrote Excel daily matrix: {path}")

def main():
    ensure_dirs()
    git_fetch_all()

    target_date = parse_target_date_ist()
    yyyymmdd = target_date.strftime("%Y%m%d")
    deadline_ist = ist_deadline_for_date(target_date)

    include_main_fallback = env_flag("INCLUDE_MAIN_FALLBACK", False)
    enforce_single_file = env_flag("ENFORCE_SINGLE_FILE", False)
    main_branch = env_str("MAIN_BRANCH_NAME", "main")

    print(f"Generating report for {target_date.isoformat()} IST (deadline {deadline_ist.strftime('%Y-%m-%d %H:%M:%S %Z')})")
    print(f"- INCLUDE_MAIN_FALLBACK = {include_main_fallback}")
    print(f"- ENFORCE_SINGLE_FILE   = {enforce_single_file}")
    print(f"- MAIN_BRANCH_NAME      = {main_branch}")

    students = read_students()

    submitted_rows: list[tuple[str, str, str, str]] = []   # (username, full_name, file_name, commit_time_ist)
    late_rows: list[tuple[str, str, str, str]] = []        # (username, full_name, file_name, commit_time_ist)
    not_submitted_rows: list[tuple[str, str]] = []         # (username, full_name)
    naming_issues: list[tuple[str, str, str, str]] = []    # (username, full_name, issue, details)

    for s in students:
        username = s["github_username"]
        full_name = s.get("full_name", "")

        ref, _ = preferred_ref_for_student(username)
        fallback_mode = False

        if ref is None:
            if include_main_fallback and ref_exists(main_branch):
                ref = main_branch
                fallback_mode = True
                verbose(f"{username}: no branch found; USING main fallback")
            else:
                not_submitted_rows.append((username, full_name))
                verbose(f"{username}: no branch or remote-tracking branch found")
                continue

        files = list_branch_files(ref, SUBMISSIONS_DIR)

        # Collect all matching files for the target date
        matches = []
        for fn in files:
            m = FILENAME_DATE_RE.match(fn)
            if not m:
                continue
            if m.group("yyyymmdd") == yyyymmdd:
                idx_str = m.group("idx")
                idx = int(idx_str) if idx_str is not None else 0
                matches.append((idx, fn))

        if not matches:
            not_submitted_rows.append((username, full_name))
            continue

        # Sort by index then filename for stable output
        matches.sort(key=lambda t: (t[0], t[1]))

        if enforce_single_file and len(matches) > 1:
            naming_issues.append(
                (username, full_name, "MULTIPLE_FILES", ";".join([fn for _, fn in matches]))
            )

        # Prefix sanity on each file
        for _, fn in matches:
            prefix = fn.rsplit("_", 1)[0] if fn.lower().endswith(".py") else fn
            if len(prefix) < 3:
                naming_issues.append((username, full_name, "SUSPICIOUS_PREFIX", fn))

        # For each file, compute commit time and lateness
        for _, fn in matches:
            file_path = f"{SUBMISSIONS_DIR}/{fn}"
            commit_iso = get_last_commit_iso_for_path(ref, file_path)

            is_late = False
            commit_ist_str = ""
            if commit_iso:
                try:
                    dt = datetime.fromisoformat(commit_iso.replace("Z", "+00:00"))
                    commit_ist = dt.astimezone(IST)
                    commit_ist_str = commit_ist.isoformat()
                    if commit_ist > deadline_ist:
                        is_late = True
                except Exception:
                    naming_issues.append((username, full_name, "BAD_COMMIT_TIME", commit_iso or ""))

            submitted_rows.append((username, full_name, fn, commit_ist_str))
            if is_late:
                late_rows.append((username, full_name, fn, commit_ist_str))

        if fallback_mode:
            naming_issues.append(
                (username, full_name, "MAIN_FALLBACK_USED",
                 f"Scanned on {main_branch}; commit times may reflect merge/squash time.")
            )

    # === Write per-day CSVs and Markdown ===
    submitted_csv = os.path.join(REPORTS_DIR, f"submitted_{yyyymmdd}.csv")
    not_submitted_csv = os.path.join(REPORTS_DIR, f"not_submitted_{yyyymmdd}.csv")
    late_csv = os.path.join(REPORTS_DIR, f"late_{yyyymmdd}.csv")
    naming_csv = os.path.join(REPORTS_DIR, f"naming_issues_{yyyymmdd}.csv")
    md_report = os.path.join(REPORTS_DIR, f"daily_report_{yyyymmdd}.md")

    with open(submitted_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["github_username", "full_name", "file_name", "commit_time_ist"])
        w.writerows(submitted_rows)

    with open(not_submitted_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["github_username", "full_name"])
        w.writerows(not_submitted_rows)

    with open(late_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["github_username", "full_name", "file_name", "commit_time_ist"])
        w.writerows(late_rows)

    with open(naming_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["github_username", "full_name", "issue", "details"])
        w.writerows(naming_issues)

    with open(md_report, "w", encoding="utf-8") as f:
        f.write(f"# Daily Submission Report - {target_date.isoformat()} (IST)\n\n")
        f.write(f"**Deadline:** {deadline_ist.strftime('%Y-%m-%d %H:%M:%S %Z')}\n\n")
        f.write(f"## Summary\n")
        f.write(f"- Total students: **{len(students)}**\n")
        submitted_students = sorted({u for (u, *_rest) in submitted_rows})
        f.write(f"- Submitted (students): **{len(submitted_students)}**\n")
        f.write(f"- Total files submitted: **{len(submitted_rows)}**\n")
        f.write(f"- Not submitted (students): **{len(not_submitted_rows)}**\n")
        f.write(f"- Late submissions (files): **{len(late_rows)}**\n")
        f.write(f"- Naming / warnings: **{len(naming_issues)}**\n\n")

        f.write("## Submitted Files\n")
        if submitted_rows:
            for u, name, fn, ctime in sorted(submitted_rows, key=lambda r: (r[0], r[2])):
                label = f"{u} ({name})" if name else u
                f.write(f"- {label} — `{fn}`{f' — {ctime}' if ctime else ''}\n")
        else:
            f.write("- None\n")

        f.write("\n## Not Submitted (Students)\n")
        if not_submitted_rows:
            for u, name in sorted(not_submitted_rows):
                label = f"{u} ({name})" if name else u
                f.write(f"- {label}\n")
        else:
            f.write("- None\n")

        f.write("\n## Late Submissions (Files)\n")
        if late_rows:
            for u, name, fn, ctime in sorted(late_rows, key=lambda r: (r[0], r[2])):
                label = f"{u} ({name})" if name else u
                f.write(f"- {label} — `{fn}` — {ctime}\n")
        else:
            f.write("- None\n")

        f.write("\n## Naming / Duplicate / Warnings\n")
        if naming_issues:
            for u, name, issue, details in sorted(naming_issues):
                label = f"{u} ({name})" if name else u
                f.write(f"- {label} — **{issue}** — {details}\n")
        else:
            f.write("- None\n")

    print(f"Generated reports for {target_date.isoformat()} IST:")
    print(f"- {md_report}")
    print(f"- {submitted_csv}")
    print(f"- {not_submitted_csv}")
    print(f"- {late_csv}")
    print(f"- {naming_csv}")

    # === Build & write Excel daily matrix ===
    dates, date_to_counts = build_daily_counts_from_reports(REPORTS_DIR)
    excel_path = os.path.join(REPORTS_DIR, "submissions_daily_matrix.xlsx")
    write_excel_daily_matrix(excel_path, students, dates, date_to_counts)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)